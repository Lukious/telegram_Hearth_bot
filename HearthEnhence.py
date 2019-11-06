import time
import telepot
from datetime import datetime
from telepot.loop import MessageLoop
import requests
from  bs4 import BeautifulSoup
import sys
import openpyxl

wb = openpyxl.load_workbook('.\\Hearth_bot_data.xlsx')
first_sheet = wb.get_sheet_names()[0]
worksheet = wb.get_sheet_by_name('총합')

bot = telepot.Bot('#keyarea')


#카드 DB 업데이트 이후 정상작동 확인용(시작할때 리소스 조금먹는 애니까 건들지 말자!)
for row in range(2,worksheet.max_row+1):  
    for column in "B":  
        cell_name = "{}{}".format(column, row)
        worksheet[cell_name].value 
        print(worksheet[cell_name].value)



def handle(msg):
    content_type, chat_type, chat_id = telepot.glance(msg)
    print(content_type, chat_type, chat_id)

    if content_type == 'text':
        text = msg['text']
        mode = text [:5]
        inputed_data = text[6:]
        print(msg['text'])

        ##이름검색자 시작.
        if '/name' in mode:
            worksheet = wb.get_sheet_by_name('총합')
            for row in range(2,worksheet.max_row+1):  
                for column in "B":  
                    cell_name = "{}{}".format(column, row)
                    worksheet[cell_name].value #여기 잠깐 고민좀(39-41)
            count = 0
            only_number = -1
            choice = []
            jchoice = '존재하지 않는 카드명 입니다.'
            for number in range(2,worksheet.max_row+1):
               	cell = "{}{}".format("B",number)
                if(inputed_data and (worksheet[cell].value).find(inputed_data) != -1):
                    print(worksheet[cell].value)
                    choice.append(str(worksheet[cell].value) + str("\n"))
                    count=count+1
                    only_number = number
                    jchoice = ''.join(choice)

            bot.sendMessage(chat_id, jchoice)

                    ##print((worksheet[cell_name].value).find(msg['text']))


            if(count == 1):
            	space = []
            	for alphabet in ['A','B','C','D','E','F','G','H','I','J']:
            		c = "{}{}".format(alphabet,only_number)
            		p = "{}{}".format(alphabet,1)
            		space.append(str(worksheet[p].value) + str("\t: ") + str(worksheet[c].value) + str("\n"))
            	jpnt = ''.join(space)
            	print(jpnt)
            	bot.sendMessage(chat_id, jpnt)


        #덱 검색자 시작.
        if '/deck' in mode:
            worksheet = wb.get_sheet_by_name('덱리')
            for row in range(2,worksheet.max_row+1):  
                for column in "A":  
                    cell_name = "{}{}".format(column, row)
                    worksheet[cell_name].value #여기 잠깐 고민좀(39-41)
            count = 0
            only_number = -1
            choice = []
            jchoice = '존재하지 않는 덱이름 입니다.'
            for number in range(2,worksheet.max_row+1):
                cell = "{}{}".format("A",number)
                if((worksheet[cell].value).find(inputed_data) != -1):
                    print(worksheet[cell].value)
                    choice.append(str(worksheet[cell].value) + str("\n"))
                    count=count+1
                    only_number = number
                    jchoice = ''.join(choice)

            bot.sendMessage(chat_id, jchoice)

                    ##print((worksheet[cell_name].value).find(msg['text']))


            if(count == 1):
                space = []
                for alphabet in ['B']:
                    c = "{}{}".format(alphabet,only_number)
                    p = "{}{}".format(alphabet,1)
                    space.append(str(worksheet[c].value) + str("\n"))
                jpnt = ''.join(space)
                print(jpnt)
                bot.sendMessage(chat_id, jpnt)



        #친선퀘 도둑 검색자 시작.
        if '/foxs' in mode:
            worksheet = wb.get_sheet_by_name('도둑')
            for row in range(2,worksheet.max_row+1):  
                for column in "A":  
                    cell_name = "{}{}".format(column, row)
                    worksheet[cell_name].value #여기 잠깐 고민좀(39-41)
            count = 0
            only_number = -1
            choice = []
            jchoice = '블랙리스트에 없습니다.'
            for number in range(2,worksheet.max_row+1):
                cell = "{}{}".format("A",number)

                if(inputed_data and (worksheet[cell].value).find(inputed_data) != -1):
                    print(worksheet[cell].value)
                    choice.append(str("\'")+str(inputed_data) + str("\'") + str("님이 블랙리스트에 있습니다!\n"))
                    count=count+1
                    only_number = number
                    jchoice = ''.join(choice)

            bot.sendMessage(chat_id, jchoice)




bot.message_loop(handle)




# Keep the program running.
while 1:
    time.sleep(10)

##2017-11-07 최신버전
